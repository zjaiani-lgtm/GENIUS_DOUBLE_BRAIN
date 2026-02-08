# execution/signal_generator.py
import os
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional

import ccxt
import openpyxl

from execution.signal_client import append_signal
from execution.db.repository import get_open_positions_count

EXCEL_PATH = Path(os.getenv("BRAIN_XLSX_PATH", "/var/data/brain.xlsx"))


def _bool(v: Any) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y")


def _read_kv_sheet(ws) -> Dict[str, Any]:
    """
    Reads GENERATOR_CONFIG formatted as:
      Column A = key
      Column B = value
    """
    cfg: Dict[str, Any] = {}
    r = 1
    while True:
        k = ws[f"A{r}"].value
        v = ws[f"B{r}"].value
        if k is None:
            break
        cfg[str(k).strip()] = v
        r += 1
        if r > 200:
            break
    return cfg


def _parse_symbols(cfg: Dict[str, Any]) -> List[str]:
    symbols_csv = str(cfg.get("SYMBOLS_CSV") or "").strip()
    if symbols_csv:
        return [s.strip() for s in symbols_csv.split(",") if s.strip()]
    # fallback single symbol
    sym = str(cfg.get("SYMBOL") or "BTC/USDT").strip()
    return [sym]


def _safe_float(v: Any, default: float) -> float:
    try:
        return float(v)
    except Exception:
        return default


def _safe_int(v: Any, default: int) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _load_symbol_overrides(wb) -> Dict[str, Dict[str, float]]:
    """
    Optional sheet: SYMBOL_OVERRIDES
    Columns: SYMBOL | USDT_SIZE_OVERRIDE | TP_PCT_OVERRIDE | SL_PCT_OVERRIDE
    """
    if "SYMBOL_OVERRIDES" not in wb.sheetnames:
        return {}

    ws = wb["SYMBOL_OVERRIDES"]
    overrides: Dict[str, Dict[str, float]] = {}

    # header at row 1
    for r in range(2, 500):
        sym = ws[f"A{r}"].value
        if sym is None:
            break
        sym = str(sym).strip()
        if not sym:
            continue

        o: Dict[str, float] = {}
        usdt = ws[f"B{r}"].value
        tp = ws[f"C{r}"].value
        sl = ws[f"D{r}"].value

        if usdt not in (None, ""):
            o["USDT_SIZE"] = _safe_float(usdt, 0.0)
        if tp not in (None, ""):
            o["TP_PCT"] = _safe_float(tp, 0.0)
        if sl not in (None, ""):
            o["SL_PCT"] = _safe_float(sl, 0.0)

        if o:
            overrides[sym] = o

    return overrides


def _get_last_signal_time_map() -> Dict[str, float]:
    """
    Very small in-process cache in env-less runtime: use a module-level dict by attaching to function.
    """
    if not hasattr(_get_last_signal_time_map, "_cache"):
        _get_last_signal_time_map._cache = {}  # type: ignore
    return _get_last_signal_time_map._cache  # type: ignore


def _sma(values: List[float], n: int) -> Optional[float]:
    if len(values) < n:
        return None
    return sum(values[-n:]) / n


def run_once(outbox_path: str) -> bool:
    """
    Reads brain.xlsx (brain_FINAL layout), scans multiple symbols, and writes at most ONE signal per loop.
    Returns True if a signal was written.
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"brain.xlsx not found at {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "GENERATOR_CONFIG" not in wb.sheetnames:
        raise ValueError("GENERATOR_CONFIG sheet missing")

    cfg = _read_kv_sheet(wb["GENERATOR_CONFIG"])
    enabled = _bool(cfg.get("ENABLED", False))
    if not enabled:
        return False

    max_open = _safe_int(cfg.get("MAX_OPEN_POSITIONS"), 1)
    if get_open_positions_count() >= max_open:
        return False

    tf = str(cfg.get("TIMEFRAME") or "1m").strip()
    limit = _safe_int(cfg.get("LIMIT"), 50)
    ma_period = _safe_int(cfg.get("MA_PERIOD"), 20)
    min_conf = _safe_float(cfg.get("MIN_CONF"), 0.70)

    usdt_size_default = _safe_float(cfg.get("USDT_SIZE"), 1.0)
    tp_pct_default = _safe_float(cfg.get("TP_PCT"), 0.03)
    sl_pct_default = _safe_float(cfg.get("SL_PCT"), 0.015)

    sl_buf = _safe_float(cfg.get("SL_LIMIT_BUFFER_PCT"), 0.001)
    cooldown_s = _safe_int(cfg.get("COOLDOWN_SECONDS"), 600)

    direction = str(cfg.get("DIRECTION") or "LONG").strip().upper()
    if direction != "LONG":
        # spot only in this demo
        return False

    symbols = _parse_symbols(cfg)
    overrides = _load_symbol_overrides(wb)

    ex = ccxt.binance({"enableRateLimit": True, "options": {"defaultType": "spot"}})

    last_map = _get_last_signal_time_map()
    now_ts = datetime.now(timezone.utc).timestamp()

    # scan symbols; create at most ONE signal
    for symbol in symbols:
        # cooldown per symbol
        last_ts = float(last_map.get(symbol, 0.0))
        if now_ts - last_ts < cooldown_s:
            continue

        # per-symbol overrides (optional)
        o = overrides.get(symbol, {})
        usdt_size = float(o.get("USDT_SIZE", usdt_size_default))
        tp_pct = float(o.get("TP_PCT", tp_pct_default))
        sl_pct = float(o.get("SL_PCT", sl_pct_default))

        ohlcv = ex.fetch_ohlcv(symbol, timeframe=tf, limit=limit)
        closes = [c[4] for c in ohlcv]
        last = closes[-1]
        ma = _sma(closes, ma_period)
        if ma is None:
            continue

        # Example rule: BUY if price > MA
        confidence = 0.75 if last > ma else 0.50
        if not (last > ma and confidence >= min_conf):
            continue

        position_size = usdt_size / last

        tp_price = last * (1 + tp_pct)
        sl_stop = last * (1 - sl_pct)
        sl_limit = sl_stop * (1 - sl_buf)

        signal = {
            "signal_id": f"DYZEN-{uuid.uuid4().hex[:12]}",
            "created_at_utc": _utc_now(),
            "final_verdict": "TRADE",
            "certified_signal": True,
            "execution": {
                "symbol": symbol,
                "direction": "LONG",
                "position_size": float(position_size),
                "entry": {"type": "MARKET"},
                "exits": {
                    "tp": {"type": "LIMIT", "price": float(tp_price)},
                    "sl": {"type": "STOP_LIMIT", "stop_price": float(sl_stop), "limit_price": float(sl_limit)},
                },
            },
            "meta": {
                "tf": tf,
                "last": float(last),
                "ma": float(ma),
                "confidence": float(confidence),
            },
        }

        append_signal(signal, outbox_path)

        # mark cooldown
        last_map[symbol] = now_ts
        return True

    return False
