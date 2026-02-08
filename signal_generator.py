import os
import time
import json
import uuid
import shutil
from datetime import datetime, timezone
from pathlib import Path

import ccxt
import openpyxl

# Disk paths on Render
OUTBOX_PATH = Path(os.getenv("SIGNAL_OUTBOX_PATH", "/var/data/signal_outbox.json"))
EXCEL_PATH = Path(os.getenv("BRAIN_XLSX_PATH", "/var/data/brain.xlsx"))

# Bundled fallback in repo (so you can deploy without manually uploading first)
BUNDLED_EXCEL = Path(os.getenv("BUNDLED_BRAIN_XLSX", "assets/brain.xlsx"))

SLEEP_S = float(os.getenv("GEN_LOOP_SECONDS", "10"))


def ensure_outbox():
    OUTBOX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not OUTBOX_PATH.exists() or OUTBOX_PATH.stat().st_size == 0:
        OUTBOX_PATH.write_text(json.dumps({"signals": []}, indent=2), encoding="utf-8")


def load_outbox():
    ensure_outbox()
    return json.loads(OUTBOX_PATH.read_text(encoding="utf-8"))


def save_outbox(payload):
    tmp = OUTBOX_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(OUTBOX_PATH)


def ensure_excel():
    """
    Prefer /var/data/brain.xlsx (editable on disk),
    but if missing, copy from repo assets/brain.xlsx once.
    """
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if EXCEL_PATH.exists() and EXCEL_PATH.stat().st_size > 0:
        return

    if BUNDLED_EXCEL.exists():
        shutil.copyfile(BUNDLED_EXCEL, EXCEL_PATH)
        print(f"[GEN] brain.xlsx missing -> copied from {BUNDLED_EXCEL} to {EXCEL_PATH}")
        return

    raise FileNotFoundError(f"brain.xlsx not found at {EXCEL_PATH} and no bundled file at {BUNDLED_EXCEL}")


def read_config():
    ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["GENERATOR_CONFIG"]

    def cell(addr, cast=str, default=None):
        v = ws[addr].value
        if v is None:
            return default
        try:
            return cast(v)
        except Exception:
            return default

    symbol = cell("B1", str, "BTC/USDT").strip()
    tf = cell("B2", str, "1m").strip()
    limit = cell("B3", int, 50)
    ma_period = cell("B4", int, 20)
    min_conf = cell("B5", float, 0.70)
    usdt_size = cell("B6", float, 5.0)
    tp_pct = cell("B7", float, 0.03)
    sl_pct = cell("B8", float, 0.015)

    enabled_raw = ws["B9"].value
    enabled = str(enabled_raw).strip().lower() in ("true", "1", "yes", "y")

    return {
        "symbol": symbol,
        "tf": tf,
        "limit": limit,
        "ma_period": ma_period,
        "min_conf": min_conf,
        "usdt_size": usdt_size,
        "tp_pct": tp_pct,
        "sl_pct": sl_pct,
        "enabled": enabled,
    }


def sma(values, n):
    if len(values) < n:
        return None
    return sum(values[-n:]) / n


def main():
    ensure_outbox()
    ensure_excel()

    ex = ccxt.binance({
        "enableRateLimit": True,
        "options": {"defaultType": "spot"},
    })

    print(f"[GEN] starting | EXCEL={EXCEL_PATH} OUTBOX={OUTBOX_PATH} SLEEP={SLEEP_S}s")

    while True:
        try:
            cfg = read_config()
            if not cfg["enabled"]:
                print("[GEN] disabled in Excel (ENABLED=false) — waiting...")
                time.sleep(SLEEP_S)
                continue

            symbol = cfg["symbol"]
            tf = cfg["tf"]
            limit = cfg["limit"]
            ma_period = cfg["ma_period"]
            min_conf = cfg["min_conf"]
            usdt_size = cfg["usdt_size"]
            tp_pct = cfg["tp_pct"]
            sl_pct = cfg["sl_pct"]

            ohlcv = ex.fetch_ohlcv(symbol, timeframe=tf, limit=limit)
            closes = [c[4] for c in ohlcv]
            last = closes[-1]
            ma = sma(closes, ma_period)

            if ma is None:
                print("[GEN] not enough candles for MA — waiting...")
                time.sleep(SLEEP_S)
                continue

            # ✅ Example rule (მერე შენს Excel/DYZEN წესზე გადავიყვანთ):
            # BUY if price > MA
            confidence = 0.75 if last > ma else 0.50
            decision = "TRADE" if (last > ma and confidence >= min_conf) else "NO_TRADE"

            print(f"[GEN] snapshot | last={last:.2f} ma={ma:.2f} conf={confidence:.2f} decision={decision}")

            if decision != "TRADE":
                time.sleep(SLEEP_S)
                continue

            amount = usdt_size / last
            tp_price = last * (1 + tp_pct)
            sl_price = last * (1 - sl_pct)

            sig = {
                "signal_id": f"DYZEN-{uuid.uuid4().hex[:12]}",
                "created_at_utc": datetime.now(timezone.utc).isoformat(),
                "final_verdict": "TRADE",
                "symbol": symbol,
                "side": "BUY",
                "amount": float(amount),
                "tp_price": float(tp_price),
                "sl_price": float(sl_price),
                "confidence": float(confidence),
                "certified_signal": True
            }

            out = load_outbox()
            out.setdefault("signals", []).append(sig)
            save_outbox(out)

            print(f"[GEN] SIGNAL_WRITTEN | id={sig['signal_id']} symbol={symbol} amount={amount:.8f}")

        except Exception as e:
            print(f"[GEN] ERROR: {e}")

        time.sleep(SLEEP_S)


if __name__ == "__main__":
    main()
